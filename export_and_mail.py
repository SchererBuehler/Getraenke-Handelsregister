from __future__ import annotations

import argparse
import html
import json
import os
import smtplib
import ssl
from datetime import date
from email.message import EmailMessage
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


REPORTS_DIR = Path("reports")
EXPORTS_DIR = Path("exports")


def repair_mojibake(value: str) -> str:
    """
    Repariert typische UTF-8-Zeichen, die fälschlich als Latin-1
    interpretiert wurden, z. B. 'GeschÃ¤ftsfÃ¼hrer'.
    """
    if not value:
        return value

    suspicious = ("Ã", "Â", "â€", "ðŸ")

    if not any(marker in value for marker in suspicious):
        return value

    try:
        repaired = value.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return value

    return repaired


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, (list, tuple, set)):
        return ", ".join(clean_text(item) for item in value if item is not None)

    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)

    text = html.unescape(str(value))
    return repair_mojibake(text).strip()


def get_nested(data: dict[str, Any], *keys: str) -> Any:
    current: Any = data

    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)

    return current


def normalize_uid(uid: str) -> str:
    uid = "".join(character for character in uid if character.isalnum())

    if uid.startswith("CHE") and len(uid) == 12:
        return f"{uid[:3]}-{uid[3:6]}.{uid[6:9]}.{uid[9:12]}"

    return uid


def extract_row(entry: dict[str, Any]) -> list[Any]:
    company_short = get_nested(entry, "raw", "companyShort") or {}
    publication = get_nested(entry, "raw", "sogcPublication") or {}

    company_name = (
        entry.get("company_name")
        or company_short.get("name")
        or ""
    )

    uid = (
        entry.get("uid")
        or company_short.get("uid")
        or ""
    )

    canton = (
        entry.get("canton")
        or publication.get("registryOfCommerceCanton")
        or ""
    )

    locality = (
        entry.get("locality")
        or company_short.get("legalSeat")
        or ""
    )

    mutation_types = publication.get("mutationTypes") or []
    mutation_keys = [
        item.get("key", "")
        for item in mutation_types
        if isinstance(item, dict)
    ]

    legal_form = get_nested(company_short, "legalForm", "name", "de") or ""
    status = company_short.get("status", "")

    message = publication.get("message") or entry.get("purpose") or ""

    return [
        clean_text(entry.get("publication_date")),
        clean_text(entry.get("event_type")),
        clean_text(company_name),
        normalize_uid(clean_text(uid)),
        clean_text(canton),
        clean_text(locality),
        clean_text(legal_form),
        clean_text(status),
        clean_text(entry.get("categories", [])),
        entry.get("confidence", ""),
        clean_text(entry.get("contact_names", [])),
        clean_text(entry.get("emails", [])),
        clean_text(entry.get("phones", [])),
        clean_text(entry.get("website")),
        clean_text(mutation_keys),
        clean_text(message),
        clean_text(entry.get("publication_id")),
    ]


def create_excel(report_path: Path, excel_path: Path) -> int:
    with report_path.open("r", encoding="utf-8") as file:
        records = json.load(file)

    if not isinstance(records, list):
        raise ValueError(
            f"Der JSON-Bericht muss eine Liste enthalten: {report_path}"
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Treffer"

    headers = [
        "Publikationsdatum",
        "Ereignistyp",
        "Firma",
        "UID",
        "Kanton",
        "Ort",
        "Rechtsform",
        "Status",
        "Kategorien",
        "Konfidenz",
        "Kontaktpersonen",
        "E-Mail-Adressen",
        "Telefonnummern",
        "Website",
        "Mutationsarten",
        "Publikationstext",
        "Publikations-ID",
    ]

    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for record in records:
        if isinstance(record, dict):
            sheet.append(extract_row(record))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    width_by_column = {
        1: 18,
        2: 16,
        3: 36,
        4: 20,
        5: 10,
        6: 24,
        7: 34,
        8: 14,
        9: 25,
        10: 12,
        11: 35,
        12: 35,
        13: 25,
        14: 35,
        15: 30,
        16: 100,
        17: 28,
    }

    for column_number, width in width_by_column.items():
        column_letter = get_column_letter(column_number)
        sheet.column_dimensions[column_letter].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # Zweites Blatt mit einer kleinen Zusammenfassung
    summary = workbook.create_sheet("Zusammenfassung")
    summary.append(["Kennzahl", "Wert"])
    summary.append(["Anzahl Treffer", len(records)])

    event_counts: dict[str, int] = {}
    category_counts: dict[str, int] = {}

    for record in records:
        if not isinstance(record, dict):
            continue

        event_type = clean_text(record.get("event_type")) or "unbekannt"
        event_counts[event_type] = event_counts.get(event_type, 0) + 1

        categories = record.get("categories") or []
        if isinstance(categories, list):
            for category in categories:
                category_name = clean_text(category)
                if category_name:
                    category_counts[category_name] = (
                        category_counts.get(category_name, 0) + 1
                    )

    summary.append([])
    summary.append(["Ereignistyp", "Anzahl"])

    for event_type, count in sorted(event_counts.items()):
        summary.append([event_type, count])

    summary.append([])
    summary.append(["Kategorie", "Anzahl"])

    for category, count in sorted(
        category_counts.items(),
        key=lambda item: (-item[1], item[0]),
    ):
        summary.append([category, count])

    for cell in summary[1]:
        cell.font = Font(bold=True)

    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 15
    summary.freeze_panes = "A2"

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)

    return len(records)


def parse_recipients(value: str) -> list[str]:
    recipients = [
        item.strip()
        for item in value.replace(";", ",").split(",")
        if item.strip()
    ]

    if not recipients:
        raise ValueError("Es wurde kein E-Mail-Empfänger angegeben.")

    return recipients


def send_email(
    excel_path: Path,
    report_date: str,
    number_of_records: int,
) -> None:
    smtp_host = os.environ["SMTP_HOST"]
    smtp_port = int(os.environ.get("SMTP_PORT", "465"))
    smtp_security = os.environ.get("SMTP_SECURITY", "ssl").lower()

    smtp_username = os.environ["SMTP_USERNAME"]
    smtp_password = os.environ["SMTP_PASSWORD"]

    sender = os.environ.get("MAIL_FROM", smtp_username)
    recipients = parse_recipients(os.environ["MAIL_TO"])

    message = EmailMessage()
    message["Subject"] = (
        f"Getränke-Handelsregister: "
        f"{number_of_records} Treffer vom {report_date}"
    )
    message["From"] = sender
    message["To"] = ", ".join(recipients)

    message.set_content(
        f"""Guten Tag

Im Anhang befindet sich der Excel-Tagesbericht des
Getränke-Handelsregister-Bots für den {report_date}.

Anzahl Treffer: {number_of_records}

Ereignistypen:
- Neueintragungen
- Mutationen
- Schliessungen und Löschungen

Hinweis: Die Treffer und Kontaktdaten wurden automatisiert ermittelt und
sollten vor einer geschäftlichen Kontaktaufnahme geprüft werden.

Freundliche Grüsse
Getränke-Handelsregister-Bot
"""
    )

    attachment = excel_path.read_bytes()

    message.add_attachment(
        attachment,
        maintype=(
            "application"
        ),
        subtype=(
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        filename=excel_path.name,
    )

    context = ssl.create_default_context()

    if smtp_security == "ssl":
        with smtplib.SMTP_SSL(
            smtp_host,
            smtp_port,
            context=context,
            timeout=30,
        ) as smtp:
            smtp.login(smtp_username, smtp_password)
            smtp.send_message(message)

    elif smtp_security == "starttls":
        with smtplib.SMTP(
            smtp_host,
            smtp_port,
            timeout=30,
        ) as smtp:
            smtp.ehlo()
            smtp.starttls(context=context)
            smtp.ehlo()
            smtp.login(smtp_username, smtp_password)
            smtp.send_message(message)

    else:
        raise ValueError(
            "SMTP_SECURITY muss 'ssl' oder 'starttls' sein."
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="JSON-Tagesbericht als Excel exportieren und versenden."
    )
    parser.add_argument(
        "--date",
        required=True,
        help="Berichtsdatum im Format YYYY-MM-DD",
    )
    parser.add_argument(
        "--no-email",
        action="store_true",
        help="Excel erzeugen, aber keine E-Mail senden",
    )
    args = parser.parse_args()

    try:
        date.fromisoformat(args.date)
    except ValueError as exc:
        raise SystemExit(
            "--date muss das Format YYYY-MM-DD haben."
        ) from exc

    report_path = REPORTS_DIR / f"treffer-{args.date}.json"
    excel_path = EXPORTS_DIR / f"treffer-{args.date}.xlsx"

    if not report_path.exists():
        raise SystemExit(
            f"Der Tagesbericht wurde nicht gefunden: {report_path}"
        )

    number_of_records = create_excel(report_path, excel_path)

    print(
        f"Excel-Datei erstellt: {excel_path} "
        f"({number_of_records} Treffer)"
    )

    if args.no_email:
        return

    send_email(
        excel_path=excel_path,
        report_date=args.date,
        number_of_records=number_of_records,
    )

    print("E-Mail erfolgreich versendet.")


if __name__ == "__main__":
    main()
