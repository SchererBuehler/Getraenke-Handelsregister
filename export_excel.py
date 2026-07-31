#!/usr/bin/env python3
"""Combine one or more JSON reports into a formatted Excel workbook."""
from __future__ import annotations

import argparse
import html
import json
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(clean_text(item) for item in value if item not in (None, ""))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    text = html.unescape(str(value)).strip()
    if any(marker in text for marker in ("Ã", "Â", "â€")):
        try:
            text = text.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return text


def nested(data: dict[str, Any], *keys: str) -> Any:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def format_uid(value: Any) -> str:
    uid = "".join(ch for ch in clean_text(value) if ch.isalnum())
    if uid.startswith("CHE") and len(uid) == 12:
        return f"CHE-{uid[3:6]}.{uid[6:9]}.{uid[9:12]}"
    return uid


def load_reports(paths: Iterable[Path]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for path in paths:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if not isinstance(payload, list):
            raise ValueError(f"JSON-Bericht muss eine Liste enthalten: {path}")
        for entry in payload:
            if not isinstance(entry, dict):
                continue
            publication_id = clean_text(entry.get("publication_id"))
            key = publication_id or json.dumps(entry, sort_keys=True, ensure_ascii=False)
            if key not in seen:
                seen.add(key)
                records.append(entry)
    return records


def row_for(entry: dict[str, Any]) -> list[Any]:
    company = nested(entry, "raw", "companyShort") or {}
    publication = nested(entry, "raw", "sogcPublication") or {}
    mutations = publication.get("mutationTypes") or []
    mutation_keys = [item.get("key", "") for item in mutations if isinstance(item, dict)]
    sources = entry.get("contact_sources") or {}
    return [
        clean_text(entry.get("publication_date")),
        clean_text(entry.get("event_type")),
        clean_text(entry.get("company_name") or company.get("name")),
        format_uid(entry.get("uid") or company.get("uid")),
        clean_text(entry.get("canton") or publication.get("registryOfCommerceCanton")),
        clean_text(entry.get("locality") or company.get("legalSeat")),
        clean_text(nested(company, "legalForm", "name", "de")),
        clean_text(company.get("status")),
        clean_text(entry.get("categories", [])),
        entry.get("confidence", ""),
        clean_text(entry.get("contact_names", [])),
        clean_text(entry.get("emails", [])),
        clean_text(entry.get("phones", [])),
        clean_text(entry.get("website")),
        clean_text(mutation_keys),
        clean_text(publication.get("message") or entry.get("purpose")),
        clean_text(sources),
        clean_text(entry.get("source_url")),
        clean_text(entry.get("publication_id")),
    ]


def create_workbook(records: list[dict[str, Any]], output: Path, period: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Treffer"
    headers = [
        "Publikationsdatum", "Ereignistyp", "Firma", "UID", "Kanton", "Ort",
        "Rechtsform", "Status", "Kategorien", "Konfidenz", "Kontaktpersonen",
        "E-Mail-Adressen", "Telefonnummern", "Website", "Mutationsarten",
        "Publikationstext", "Kontaktquellen", "Quellen-URL", "Publikations-ID",
    ]
    ws.append(headers)
    for record in sorted(records, key=lambda r: (clean_text(r.get("publication_date")), clean_text(r.get("company_name")))):
        ws.append(row_for(record))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [16, 15, 34, 20, 9, 22, 30, 13, 24, 12, 32, 32, 24, 32, 28, 80, 40, 35, 28]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if ws.max_row >= 2:
        table = Table(displayName="TrefferTabelle", ref=f"A1:S{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)

    summary = wb.create_sheet("Zusammenfassung")
    summary.append(["Wochenbericht", period])
    summary.append(["Anzahl Treffer", len(records)])
    summary.append([])
    summary.append(["Ereignistyp", "Anzahl"])
    event_counts = Counter(clean_text(r.get("event_type")) or "unbekannt" for r in records)
    for name, count in sorted(event_counts.items()):
        summary.append([name, count])
    summary.append([])
    summary.append(["Kategorie", "Anzahl"])
    category_counts: Counter[str] = Counter()
    for record in records:
        categories = record.get("categories") or []
        if isinstance(categories, list):
            category_counts.update(clean_text(category) for category in categories if clean_text(category))
    for name, count in category_counts.most_common():
        summary.append([name, count])
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 18
    for row_number in (1, 4, 4 + len(event_counts) + 3):
        for cell in summary[row_number]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="JSON-Berichte als Excel-Datei exportieren")
    parser.add_argument("--input", nargs="+", required=True, help="Eine oder mehrere JSON-Dateien")
    parser.add_argument("--output", required=True, help="Zielpfad der XLSX-Datei")
    parser.add_argument("--period", default="", help="Bezeichnung des Berichtszeitraums")
    args = parser.parse_args()
    paths = [Path(value) for value in args.input]
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise SystemExit(f"Fehlende JSON-Berichte: {', '.join(missing)}")
    records = load_reports(paths)
    create_workbook(records, Path(args.output), args.period)
    print(f"Excel-Datei erstellt: {args.output} ({len(records)} Treffer)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
